import openpyxl as xl
wb = xl.load_workbook("inventory.xlsx")
wb_sheet = wb["home"]
#products per supplier
#total inventory price per supplier
#product with inventory less than 10
#add a new column total inventory price
#supplier with the highest number of inventory
product_per_sup = {}
prod_10 = {}
total_per_sup = {}
for i in range(2,wb_sheet.max_row+1):
    suplier_name = wb_sheet.cell(i, 4).value
    price = wb_sheet.cell(i, 3).value
    inventory = wb_sheet.cell(i, 2).value
    product = wb_sheet.cell(i, 1).value
    total_price = price*inventory
    new_total_inve_price_col = wb_sheet.cell(i, 5)
    if suplier_name in total_per_sup:
        total_per_sup[suplier_name] = total_per_sup[suplier_name]+total_price
    else:
        total_per_sup[suplier_name] = total_price
    #products per supplier
    if suplier_name in product_per_sup:
        product_per_sup[suplier_name] += 1
    else:
        product_per_sup[suplier_name] = 1
    if inventory < 10:
        prod_10[product] = inventory
    new_total_inve_price_col.value = total_price
        
print(product_per_sup)
print(prod_10)
print(total_per_sup)
wb.save("inventory_updated.xlsx")
